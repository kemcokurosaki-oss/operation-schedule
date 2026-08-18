# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    (1, "8/17", "新機能／UI改善",
     "完了工事一覧の詳細画面が「操業」「出張」の2タブから、通常画面と同じ「計画」「社内試運転」「出張」の3タブ構成に変更されました。完了した工事でも計画段階のタスクだけを分けて確認できるようになりました。",
     "archive.js, index.html"),
    (2, "8/17", "その他（整理）",
     "使い方ガイド（？ボタン）から、廃止された工事番号・機械・ユニット・担当者フィルターボタンの説明が削除されました（列見出しフィルターへの統合に伴う整理）。",
     "auth.js"),
    (3, "8/18", "新機能",
     "画面上部にあった「工事番号」「機械」「ユニット」「担当者」の絞り込みボタンが廃止され、代わりに表の各列見出しにExcelのオートフィルターのような▼ボタンが付きました。▼を押すと、その列に実際に表示されている値の一覧からチェックで絞り込みができ、昇順・降順の並べ替えも選べます。開始日・完了予定日の列では年→月→日のツリー形式で選べます。絞り込み中の列は▼が漏斗マークに変わって一目で分かります。",
     "index.html, data.js, gantt-setup.js, style.css"),
    (4, "8/18", "新機能",
     "「担当別」表示画面で、担当者ごとの業務バーを直接ドラッグして期間を移動したり、バーの端をつまんで開始日・終了日を伸縮できるようになりました。ダブルクリックで詳細編集画面を開くことも可能です。",
     "resource.js, style.css, gantt-setup.js"),
    (5, "8/18", "UI改善",
     "ガントチャートのバーや「担当別」画面のバーに、これまではタスク名だけだった表示が「工事番号 機械名 タスク名」とまとめて表示されるようになり、どの工事・機械の作業か一目で分かるようになりました。",
     "gantt-setup.js, resource.js"),
    (6, "8/18", "不具合修正",
     "表の「開始日」または「終了日」だけをその場で編集したときに、編集していないもう片方の日付まで勝手に変わってしまう不具合を修正しました。",
     "gantt-setup.js"),
    (7, "8/18", "不具合修正",
     "「担当別」画面をフルスクリーン表示しているときに、スクロール位置がずれたりバーの座標がおかしくなって消えてしまう不具合を修正しました。",
     "resource.js, gantt-setup.js"),
    (8, "8/18", "不具合修正",
     "「担当別」フルスクリーン画面のタスクバーをダブルクリックしても編集画面が開かなかった不具合を修正し、担当別画面からも直接タスクを編集できるようになりました。",
     "gantt-setup.js"),
    (9, "8/18", "不具合修正",
     "「＋新規タスク追加」ボタンを、工事番号を1件も選択していない状態でも、既存行の「＋」から使えば追加できるように修正しました（以前は工事番号を選んでいないと使えませんでした）。",
     "gantt-setup.js"),
    (10, "8/18", "新機能",
     "表の一部の列（ユニット・タスク名、出張モードでは客先・工事名）の幅を、境界線をドラッグして自由に変更できるようになりました。ダブルクリックすると内容に合わせて自動調整され、変更した幅は次回開いたときも保持されます。",
     "gantt-setup.js"),
    (11, "8/18", "仕様変更（統合）",
     "「🏭 現地試運転」という独立した表示モードが廃止され、「出張」モードに統合されました。出張タスクの中で「現地試運転」「現地SV」「調査」をプルダウンで選ぶ形になり、表示モードのボタンが1つ減ってシンプルになりました。またヘッダーの「📋 プログラム作成」ボタンの表記が「📋 計画」に変更されました。",
     "index.html, data.js, gantt-setup.js"),
    (12, "8/18", "仕様変更",
     "表示モード（計画／社内試運転／出張）を切り替えるボタンについて、選択中のボタンをもう一度押しても「担当別」画面には戻らなくなりました（担当別画面に戻るには別の「👤担当別」ボタンを使う仕様に変更）。",
     "data.js"),
    (13, "8/18", "UI改善",
     "担当者プルダウンの選択肢に「未定」が追加され、担当者が決まっていないタスクにも登録しやすくなりました。",
     "gantt-setup.js"),
    (14, "8/18", "UI改善",
     "「担当別」画面の担当者ごとの明細（タスク一覧）が、これまでの「計画→試運転→出張」という種類順の並びから、開始日が早い順の並びに変更されました（日付未定のタスクは末尾）。また各行に機械名も表示されるようになりました。",
     "resource.js"),
    (15, "8/18", "UI改善",
     "タスクの編集画面（ライトボックス）が全体的に大きく見やすくなりました（幅・文字サイズ・入力欄の高さを拡大）。",
     "style.css, gantt-setup.js"),
    (16, "8/18", "UI改善",
     "画面上部のボタン配置が整理され、「リソース表示」「＋新規タスク追加」「選択削除」「完了工事一覧」ボタンが表示モード切り替えボタンと同じ上段の行にまとめられました。またボタン群の間に区切り線が追加されました。",
     "index.html, style.css"),
    (17, "8/18", "用語統一",
     "「担当別」画面などの表記で「試運転」が「社内試運転」に統一され、他の画面と表現が揃いました。",
     "resource.js, archive.js"),
]

CATEGORY_COLORS = {
    "新機能": "DDEBF7",
    "新機能／UI改善": "DDEBF7",
    "不具合修正": "FCE4E4",
    "UI改善": "E2EFDA",
    "仕様変更": "FFF2CC",
    "仕様変更（統合）": "FFF2CC",
    "その他（整理）": "F2F2F2",
    "用語統一": "F2F2F2",
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "修正内容報告"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# タイトル
ws.merge_cells("A1:E1")
ws["A1"] = "操業工程表アプリ 修正内容報告（対象期間：2026年8月17日～8月18日）"
ws["A1"].font = Font(size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

ws["A2"] = "作成日：2026年8月18日"
ws["A2"].font = Font(size=10, italic=True, color="666666")

# ヘッダー行
headers = ["No.", "日付", "カテゴリ", "変更内容", "関連ファイル"]
header_row = 4
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

# データ行
r = header_row + 1
for no, date, cat, content, files in rows:
    ws.cell(row=r, column=1, value=no).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r, column=2, value=date).alignment = Alignment(horizontal="center", vertical="top")
    cat_cell = ws.cell(row=r, column=3, value=cat)
    cat_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    fill_color = CATEGORY_COLORS.get(cat, "FFFFFF")
    cat_cell.fill = PatternFill("solid", fgColor=fill_color)
    content_cell = ws.cell(row=r, column=4, value=content)
    content_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    files_cell = ws.cell(row=r, column=5, value=files)
    files_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = border
    ws.row_dimensions[r].height = 60
    r += 1

# 列幅
widths = {"A": 6, "B": 8, "C": 16, "D": 90, "E": 34}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = f"A{header_row+1}"

# 集計シート
summary = {}
for _, _, cat, _, _ in rows:
    key = cat.split("（")[0]
    summary[key] = summary.get(key, 0) + 1

ws2 = wb.create_sheet("集計")
ws2["A1"] = "カテゴリ別 件数"
ws2["A1"].font = Font(size=12, bold=True)
ws2["A3"] = "カテゴリ"
ws2["B3"] = "件数"
ws2["A3"].font = Font(bold=True)
ws2["B3"].font = Font(bold=True)
for i, (k, v) in enumerate(summary.items(), start=4):
    ws2.cell(row=i, column=1, value=k)
    ws2.cell(row=i, column=2, value=v)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 10

out_path = r"c:\Users\kurosaki\OneDrive - 日下部電機\デスクトップ\工程表作成\操業工程表\修正内容報告_20260817-18.xlsx"
wb.save(out_path)
print("saved:", out_path)
