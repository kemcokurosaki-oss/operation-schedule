# -*- coding: utf-8 -*-
"""4名（堀尾・本郷・黒見・野津）向けの権限変更内容 Excel を作成する使い捨てスクリプト"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "権限変更内容"

TITLE = "操業工程表 権限変更のお知らせ（堀尾・本郷・黒見・野津 の4名対象）"
SUBTITLE = "常務・黒崎・森村の3名は全モードで従来通り編集可能です。対象4名のみ、社内試運転モードでの操作範囲を below の通り変更します。"

thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_fill = PatternFill("solid", fgColor="2F5496")
header_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(bold=True, size=14, color="1F3864")
subtitle_font = Font(size=10, italic=True, color="595959")

ok_font = Font(color="0B6B0B", bold=True)
ng_font = Font(color="C00000", bold=True)
ok_fill = PatternFill("solid", fgColor="E2EFDA")
ng_fill = PatternFill("solid", fgColor="FCE4E4")
partial_fill = PatternFill("solid", fgColor="FFF2CC")
partial_font = Font(color="7F6000", bold=True)

# ---- タイトル行 ----
ws.merge_cells("A1:E1")
ws["A1"] = TITLE
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:E2")
ws["A2"] = SUBTITLE
ws["A2"].font = subtitle_font
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

# ---- 表ヘッダー ----
headers = ["操作", "計画モード", "社内試運転モード", "出張モード", "備考"]
header_row = 4
for col, text in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=text)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

rows = [
    ("担当者欄の選択・変更", "OK", "OK", "OK", ""),
    ("メモ欄の入力", "OK", "OK", "OK", ""),
    ("上記以外のセル入力\n（工事番号・機械・ユニット・タスク名・進捗・開始日・終了日 等）", "OK", "NG", "OK", "社内試運転モードのみ担当・メモ以外は入力不可"),
    ("タスクバーのドラッグ（開始日・終了日の変更）", "OK", "NG", "OK", ""),
    ("タスクバーのダブルクリック編集（詳細編集ダイアログ）", "OK", "NG", "OK", "担当・メモ以外の項目も含む全項目編集画面のため不可"),
    ("新規タスク追加", "OK", "NG", "OK", "＋ボタンも非表示になります"),
    ("タスク削除（単体・複数選択）", "OK", "NG", "OK", ""),
    ("複数行の一括編集", "OK", "NG", "OK", ""),
    ("コピー＆貼り付け", "OK", "NG", "OK", ""),
    ("右クリックメニュー（コピー・編集・削除）", "OK", "NG", "OK", "社内試運転モードの行では右クリックメニュー自体が表示されません"),
    ("担当別パネルでのタスクバードラッグ", "OK", "NG", "OK", ""),
    ("出図希望日（▼マーク）のドラッグ変更", "OK", "NG", "OK", ""),
]

for r, (op, plan, drawing, trip, note) in enumerate(rows, start=header_row + 1):
    ws.cell(row=r, column=1, value=op).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col, val in zip((2, 3, 4), (plan, drawing, trip)):
        cell = ws.cell(row=r, column=col, value=("可" if val == "OK" else "不可"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if val == "OK":
            cell.font = ok_font
            cell.fill = ok_fill
        else:
            cell.font = ng_font
            cell.fill = ng_fill
    ws.cell(row=r, column=5, value=note).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = border

# 担当・メモ行だけ強調（社内試運転モードでも可のまま）
for r in (header_row + 1, header_row + 2):
    ws.cell(row=r, column=3).fill = partial_fill
    ws.cell(row=r, column=3).font = partial_font
    ws.cell(row=r, column=3).value = "可（唯一の許可項目）"

# 列幅
widths = {"A": 46, "B": 14, "C": 20, "D": 14, "E": 46}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

for r in range(header_row, header_row + 1 + len(rows)):
    ws.row_dimensions[r].height = 34

ws.freeze_panes = "A5"

# 末尾に対象者・適用日の注記
note_row = header_row + len(rows) + 2
ws.merge_cells(f"A{note_row}:E{note_row}")
ws[f"A{note_row}"] = "対象者: 堀尾・本郷・黒見・野津　／　適用範囲: 社内試運転モード（task_type = operation）のタスクのみ。計画・出張モードのタスクは影響ありません。"
ws[f"A{note_row}"].font = Font(size=10, color="595959")
ws[f"A{note_row}"].alignment = Alignment(horizontal="left", wrap_text=True)

out_path = "操業工程表_権限変更内容(堀尾本郷黒見野津).xlsx"
wb.save(out_path)
print("saved:", out_path)
